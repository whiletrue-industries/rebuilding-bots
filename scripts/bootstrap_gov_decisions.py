"""One-time importer: Tal's Excel → Aurora.

Reads ~/Downloads/Gov_Res_fulldata_25032026.xlsx (or path passed via
--xlsx), connects to the Aurora instance pointed to by DATABASE_URL,
embeds each row's `טקסט מלא` text via OpenAI, and UPSERTs into the
documents/contexts tables for (bot=unified, context=government_decisions).

Run twice — once with DATABASE_URL pointed at staging Aurora, once at
prod. The Excel file lives only on the operator's laptop; nothing else
needs it. After bootstrap, ongoing updates come from the gov.il fetcher
(botnim/document_parser/gov_il_decisions/process.py).

Usage::

    # local dev (assumes a local Postgres with the migrations applied)
    python scripts/bootstrap_gov_decisions.py \
        --xlsx ~/Downloads/Gov_Res_fulldata_25032026.xlsx \
        --environment local

    # staging
    DATABASE_URL=postgres://... \
    OPENAI_API_KEY_STAGING=sk-... \
    python scripts/bootstrap_gov_decisions.py \
        --xlsx ~/Downloads/Gov_Res_fulldata_25032026.xlsx \
        --environment staging

The script is idempotent: it queries existing page_ids first and skips
rows already imported. So re-running after a partial failure resumes.
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import openpyxl

# Make the repo root importable when run as `python scripts/...`
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from botnim.document_parser.gov_il_decisions.aurora_writer import (  # noqa: E402
    existing_page_ids,
    get_or_create_context,
    write_decisions_batched,
)
from botnim.document_parser.gov_il_decisions.categorize import (  # noqa: E402
    ACTION_TYPES,
    DOMAINS,
)


# Number of decisions buffered before each batched flush. Each flush
# embeds all buffered chunks in one OpenAI multi-input call (capped at
# 2048 inputs per request); 500 decisions averages ~530 chunks, well
# under the cap and large enough to amortize HTTP latency.
BUFFER_SIZE = 500


logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("bootstrap_gov_decisions")


HEADERS_HE = {
    "decision_number": "מספר החלטה",
    "title": "כותרת",
    "part": "חלק",
    "effective_date": "תאריך תחולה",
    "publish_date": "תאריך פרסום",
    "government_number": "מספר ממשלה",
    "government": "ממשלה",
    "office": "משרד",
    "unit": "יחידה",
    "action_type": "סוג פעולה",
    "domain": "תחום",
    "text": "טקסט מלא",
    "has_attachment": "קובץ מצורף",
    "source_url": "קישור למקור",
    "attachment_urls": "קישורי קבצים",
}


def _str(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        iso = v.isoformat()
        return iso.split("T")[0] if "T" in iso else iso
    return str(v).strip()


def _page_id_from_url(url: str) -> str | None:
    if not url:
        return None
    return url.rstrip("/").rsplit("/", 1)[-1] or None


def _suffixed_page_id(base: str, part: str) -> str:
    """For chunked rows ('1/3', '2/3', ...), suffix with '__NofM'."""
    if not part or part == "1/1":
        return base
    n, m = part.split("/")
    return f"{base}__{n}of{m}"


def _parse_attachment_urls(raw: str) -> list[str]:
    if not raw:
        return []
    parts = [s.strip() for s in raw.replace("\n", ",").split(",")]
    return [p for p in parts if p.startswith("http")]


def bootstrap(
    *,
    xlsx_path: Path,
    environment: str,
    bot: str,
    context: str,
    buffer_size: int = BUFFER_SIZE,
) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {col: header.index(HEADERS_HE[col]) for col in HEADERS_HE}

    def _cell(row: tuple, col: str):
        # openpyxl read-only mode trims trailing empty cells, so rows where
        # the rightmost columns (e.g. attachment_urls) are empty come back
        # shorter than the header. Treat anything past the row length as
        # None / empty rather than letting IndexError abort the bootstrap.
        i = idx[col]
        return row[i] if i < len(row) else None

    cid = get_or_create_context(bot, context)
    seen = existing_page_ids(cid)
    logger.info("starting bootstrap: %d page_ids already in (%s, %s)", len(seen), bot, context)

    total = 0
    ok = 0
    skipped_existing = 0
    skipped_unknown_vocab = 0
    buffer: list[dict] = []

    def _flush() -> None:
        """Embed + write all buffered records in one batched OpenAI call."""
        nonlocal ok
        if not buffer:
            return
        try:
            result = write_decisions_batched(buffer, environment=environment)
            ok += result["decisions"]
            logger.info(
                "flushed batch: decisions=%d chunks_planned=%d chunks_written=%d",
                result["decisions"], result["chunks_planned"], result["chunks_written"],
            )
        except Exception as exc:
            # On batch failure, fall back to per-record writes so a single
            # bad record doesn't lose the rest of the buffer. Re-running
            # the script is also safe (ON CONFLICT DO NOTHING).
            logger.warning("batched flush failed (%s); retrying per-record", exc)
            for rec in buffer:
                try:
                    write_decisions_batched([rec], environment=environment)
                    ok += 1
                except Exception as inner:
                    logger.warning(
                        "write_decisions_batched failed for %s: %s",
                        rec.get("page_id"), inner,
                    )
        buffer.clear()

    for raw in rows:
        if not raw:
            continue
        total += 1

        source_url = _str(_cell(raw, "source_url"))
        base_pid = _page_id_from_url(source_url)
        if not base_pid:
            continue

        action = _str(_cell(raw, "action_type"))
        domain = _str(_cell(raw, "domain"))
        if action not in ACTION_TYPES or domain not in DOMAINS:
            skipped_unknown_vocab += 1
            continue

        part = _str(_cell(raw, "part")) or "1/1"
        page_id = _suffixed_page_id(base_pid, part)
        if page_id in seen:
            skipped_existing += 1
            continue

        title = _str(_cell(raw, "title"))
        text = _str(_cell(raw, "text"))
        metadata = {
            "decision_number": _str(_cell(raw, "decision_number")),
            "government_number": _str(_cell(raw, "government_number")),
            "government": _str(_cell(raw, "government")),
            "title": title,
            "publish_date": _str(_cell(raw, "publish_date")),
            "effective_date": _str(_cell(raw, "effective_date")),
            "office": _str(_cell(raw, "office")),
            "unit": _str(_cell(raw, "unit")),
            "action_type": action,
            "domain": domain,
            "has_attachment": bool(_cell(raw, "has_attachment")),
            "source_url": source_url,
            "attachment_urls": _parse_attachment_urls(_str(_cell(raw, "attachment_urls"))),
            "part": part,
        }

        buffer.append({
            "context_id": cid,
            "page_id": page_id,
            "title": title,
            "text": text,
            "metadata": metadata,
        })
        seen.add(page_id)

        if len(buffer) >= buffer_size:
            _flush()

        if total % 500 == 0:
            logger.info(
                "bootstrapped %d (ok=%d skipped_existing=%d skipped_unknown_vocab=%d)",
                total, ok, skipped_existing, skipped_unknown_vocab,
            )

    # Final flush of any partial buffer.
    _flush()

    logger.info(
        "DONE total=%d ok=%d skipped_existing=%d skipped_unknown_vocab=%d",
        total, ok, skipped_existing, skipped_unknown_vocab,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path.home() / "Downloads" / "Gov_Res_fulldata_25032026.xlsx",
    )
    parser.add_argument(
        "--environment",
        required=True,
        choices=["local", "staging", "production"],
    )
    parser.add_argument("--bot", default="unified")
    parser.add_argument("--context", default="government_decisions")
    parser.add_argument(
        "--buffer-size",
        type=int,
        default=BUFFER_SIZE,
        help=(
            "Decisions to buffer between batched embed+upsert flushes. "
            f"Defaults to {BUFFER_SIZE}."
        ),
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        logger.error("xlsx not found: %s", args.xlsx)
        sys.exit(2)

    bootstrap(
        xlsx_path=args.xlsx,
        environment=args.environment,
        bot=args.bot,
        context=args.context,
        buffer_size=args.buffer_size,
    )


if __name__ == "__main__":
    main()
